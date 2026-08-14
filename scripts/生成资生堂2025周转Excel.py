from __future__ import annotations

import os
import sys
from collections import defaultdict
from datetime import date
from decimal import Decimal

import pymysql
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import DB_CONFIG


BRAND = "资生堂"
PRODUCT_TYPES = ("正装", "小样")
SNAPSHOTS = (
    date(2024, 12, 31),
    date(2025, 1, 31), date(2025, 2, 28), date(2025, 3, 31),
    date(2025, 4, 30), date(2025, 5, 31), date(2025, 6, 30),
    date(2025, 7, 31), date(2025, 8, 31), date(2025, 9, 30),
    date(2025, 10, 31), date(2025, 11, 30), date(2025, 12, 31),
)
ZERO = Decimal(0)


def decimal(value: object) -> Decimal:
    return Decimal(str(value or 0))


def number(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def twelve_month_average(values: dict[date, Decimal]) -> Decimal:
    """Average the 12 monthly averages: (month opening + month ending) / 2."""
    monthly_averages = [
        (values.get(SNAPSHOTS[index - 1], ZERO) + values.get(SNAPSHOTS[index], ZERO))
        / Decimal(2)
        for index in range(1, len(SNAPSHOTS))
    ]
    return sum(monthly_averages, ZERO) / Decimal(12)


def product_key(code: object, name: object) -> str:
    normalized_code = str(code or "").strip()
    return normalized_code or f"__NAME__{str(name or '').strip()}"


def load_data() -> tuple[str, dict, list[dict], list[dict]]:
    connection = pymysql.connect(**DB_CONFIG)
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT data_version FROM ads.ads_publish_batch "
                "WHERE dataset='sales_daily' AND status='ready' "
                "ORDER BY published_at DESC LIMIT 1"
            )
            sales_version = cursor.fetchone()[0]

            cursor.execute(
                "SELECT `快照日期`, `状态` FROM ods.`历史库存快照批次` "
                "WHERE `快照日期` BETWEEN '2024-12-31' AND '2025-12-31' "
                "ORDER BY `快照日期`"
            )
            batches = cursor.fetchall()
            if tuple(row[0] for row in batches) != SNAPSHOTS or any(
                row[1] != "SUCCESS" for row in batches
            ):
                raise RuntimeError("计算2025年12个月平均库存所需的期初及月末快照不完整")

            cursor.execute(
                "SELECT `快照日期`, `货品编号`, `货品名称`, `分类`, `库存量` "
                "FROM ods.`历史库存` "
                "WHERE `快照日期` BETWEEN '2024-12-31' AND '2025-12-31' "
                "AND TRIM(`品牌`)=%s AND `分类` IN (%s,%s)",
                (BRAND, *PRODUCT_TYPES),
            )
            stock_rows = cursor.fetchall()

            cursor.execute(
                "SELECT product_code, product, product_type, quantity, paid_amount "
                "FROM ads.ads_sales_brand_turnover_item "
                "WHERE data_version=%s AND brand=%s "
                "AND sales_date>='2025-01-01' AND sales_date<'2026-01-01' "
                "AND product_type IN (%s,%s)",
                (sales_version, BRAND, *PRODUCT_TYPES),
            )
            sales_rows = cursor.fetchall()

            cursor.execute(
                "SELECT COALESCE(NULLIF(TRIM(channel),''),'未归类'), quantity, paid_amount "
                "FROM ads.ads_sales_daily_brand_channel_product "
                "WHERE data_version=%s AND brand=%s "
                "AND sales_date>='2025-01-01' AND sales_date<'2026-01-01' "
                "AND product_type IN (%s,%s)",
                (sales_version, BRAND, *PRODUCT_TYPES),
            )
            channel_rows = cursor.fetchall()
    finally:
        connection.close()

    products: dict[tuple[str, str], dict] = {}
    stock_by_type = defaultdict(lambda: defaultdict(Decimal))
    for snapshot, code, name, product_type, quantity in stock_rows:
        stock_by_type[product_type][snapshot] += decimal(quantity)
        key = (product_type, product_key(code, name))
        item = products.setdefault(
            key,
            {
                "code": str(code or "").strip(),
                "name": str(name or "").strip(),
                "type": product_type,
                "stock": defaultdict(Decimal),
                "sales": ZERO,
                "amount": ZERO,
            },
        )
        item["stock"][snapshot] += decimal(quantity)
        item["code"] = str(code or item["code"]).strip()
        item["name"] = str(name or item["name"]).strip()

    for code, name, product_type, quantity, amount in sales_rows:
        key = (product_type, product_key(code, name))
        item = products.setdefault(
            key,
            {
                "code": str(code or "").strip(),
                "name": str(name or "").strip(),
                "type": product_type,
                "stock": defaultdict(Decimal),
                "sales": ZERO,
                "amount": ZERO,
            },
        )
        item["sales"] += decimal(quantity)
        item["amount"] += decimal(amount)
        item["code"] = str(code or item["code"]).strip()
        if not item["name"]:
            item["name"] = str(name or "").strip()

    channels: dict[str, dict] = {}
    for channel, quantity, amount in channel_rows:
        item = channels.setdefault(
            str(channel), {"channel": str(channel), "sales": ZERO, "amount": ZERO}
        )
        item["sales"] += decimal(quantity)
        item["amount"] += decimal(amount)

    for item in products.values():
        item["avg_stock"] = twelve_month_average(item["stock"])
        item["ending_stock"] = item["stock"].get(SNAPSHOTS[-1], ZERO)
        item["turnover_rate"] = (
            item["sales"] / item["avg_stock"] if item["avg_stock"] > 0 else None
        )

    return sales_version, stock_by_type, list(products.values()), list(channels.values())


def build_workbook(output_path: str) -> dict:
    sales_version, stock_by_type, products, channels = load_data()
    average_stock = {
        product_type: twelve_month_average(stock_by_type[product_type])
        for product_type in PRODUCT_TYPES
    }
    sales_by_type = {
        product_type: sum(
            (item["sales"] for item in products if item["type"] == product_type), ZERO
        )
        for product_type in PRODUCT_TYPES
    }
    total_sales = sum(sales_by_type.values(), ZERO)
    total_average_stock = sum(average_stock.values(), ZERO)
    overall_rate = total_sales / total_average_stock

    channel_sales = sum((item["sales"] for item in channels), ZERO)
    if channel_sales != total_sales:
        raise RuntimeError(f"渠道与货品净销量不一致：{channel_sales} != {total_sales}")
    for item in channels:
        item["share"] = item["sales"] / total_sales if total_sales else ZERO
        item["turnover_rate"] = item["sales"] / total_average_stock
    channels.sort(key=lambda item: item["turnover_rate"], reverse=True)

    hot = sorted(
        (item for item in products if item["sales"] > 0),
        key=lambda item: (item["sales"], item["turnover_rate"] or ZERO),
        reverse=True,
    )[:50]
    slow = sorted(
        (
            item
            for item in products
            if item["ending_stock"] > 0 and item["avg_stock"] > 0
        ),
        key=lambda item: (
            0 if item["sales"] <= 0 else 1,
            item["turnover_rate"] or ZERO,
            -item["ending_stock"],
        ),
    )[:50]

    workbook = Workbook()
    workbook.remove(workbook.active)
    primary, light, white = "C65911", "FCE4D6", "FFFFFF"
    thin = Side(style="thin", color="D9E1F2")

    def setup(sheet, title: str, columns: int, note: str) -> None:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
        cell = sheet.cell(1, 1, title)
        cell.font = Font(size=16, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=primary)
        cell.alignment = Alignment(horizontal="center")
        sheet.row_dimensions[1].height = 28
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
        cell = sheet.cell(2, 1, note)
        cell.font = Font(size=10, color="595959")
        cell.fill = PatternFill("solid", fgColor=light)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[2].height = 31
        sheet.sheet_view.showGridLines = False

    def headers(sheet, labels: list[str]) -> None:
        for column, value in enumerate(labels, 1):
            cell = sheet.cell(4, column, value)
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=primary)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)

    def add_rows(
        sheet,
        rows: list[tuple],
        int_columns=(),
        decimal_columns=(),
        percent_columns=(),
        currency_columns=(),
    ) -> None:
        for row_index, row in enumerate(rows, 5):
            for column, value in enumerate(row, 1):
                cell = sheet.cell(row_index, column, value)
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="right" if isinstance(value, (int, float)) else "left",
                )
            for column in int_columns:
                sheet.cell(row_index, column).number_format = "#,##0"
            for column in decimal_columns:
                sheet.cell(row_index, column).number_format = "0.00"
            for column in percent_columns:
                sheet.cell(row_index, column).number_format = "0.00%"
            for column in currency_columns:
                sheet.cell(row_index, column).number_format = "¥#,##0.00;[Red]-¥#,##0.00"

    def finish(sheet, widths: list[int], last_row: int) -> None:
        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = f"A4:{get_column_letter(len(widths))}{last_row}"
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    summary_sheet = workbook.create_sheet("周转率汇总")
    setup(
        summary_sheet,
        "资生堂 2025年库存周转率",
        5,
        "周转率=2025年净销售数量÷2025年12个月平均库存；每月平均库存=(月初库存+月末库存)÷2，再对12个月取平均。数量口径，退货等负数销售记录计入净额。",
    )
    headers(summary_sheet, ["排序", "指标", "2025年净销售数量", "12个月平均库存", "周转率（次/年）"])
    summary_rows = [
        (1, "库存周转率（正装+小样）", number(total_sales), number(total_average_stock), number(overall_rate)),
        (2, "正装周转率", number(sales_by_type["正装"]), number(average_stock["正装"]), number(sales_by_type["正装"] / average_stock["正装"])),
        (3, "小样周转率", number(sales_by_type["小样"]), number(average_stock["小样"]), number(sales_by_type["小样"] / average_stock["小样"])),
    ]
    add_rows(summary_sheet, summary_rows, int_columns=(3, 4), decimal_columns=(5,))
    finish(summary_sheet, [9, 30, 22, 20, 20], 7)

    channel_sheet = workbook.create_sheet("渠道周转率")
    setup(
        channel_sheet,
        "资生堂 2025年渠道周转率",
        6,
        "渠道周转率=该渠道2025年正装+小样净销售数量÷资生堂整体12个月平均库存；负数表示该渠道全年退货数量超过正向销售，默认按周转率降序。",
    )
    headers(channel_sheet, ["排名", "渠道", "2025年净销售数量", "销售占比", "销售金额", "渠道周转率（次/年）"])
    channel_rows = [
        (
            rank,
            item["channel"],
            number(item["sales"]),
            number(item["share"]),
            number(item["amount"]),
            number(item["turnover_rate"]),
        )
        for rank, item in enumerate(channels, 1)
    ]
    add_rows(
        channel_sheet,
        channel_rows,
        int_columns=(3,),
        percent_columns=(4,),
        currency_columns=(5,),
        decimal_columns=(6,),
    )
    for row_index, item in enumerate(channels, 5):
        if item["sales"] < 0:
            for column in (3, 4, 5, 6):
                channel_sheet.cell(row_index, column).font = Font(color="C00000")
    finish(channel_sheet, [9, 34, 22, 15, 20, 22], 4 + len(channel_rows))

    def add_product_sheet(name: str, title: str, note: str, items: list[dict]) -> None:
        sheet = workbook.create_sheet(name)
        setup(sheet, title, 8, note)
        headers(sheet, ["排名", "货品编号", "货品名称", "分类", "2025年净销售", "12个月平均库存", "2025年末库存", "周转率（次/年）"])
        rows = [
            (
                rank, item["code"], item["name"], item["type"],
                number(item["sales"]), number(item["avg_stock"]),
                number(item["ending_stock"]), number(item["turnover_rate"]),
            )
            for rank, item in enumerate(items, 1)
        ]
        add_rows(sheet, rows, int_columns=(5, 6, 7), decimal_columns=(8,))
        finish(sheet, [9, 18, 52, 11, 17, 19, 18, 21], 4 + len(rows))

    add_product_sheet(
        "Top滞销品",
        "资生堂 2025年 Top 50 滞销品",
        "范围：2025-12-31期末库存>0且12个月平均库存>0。先排无净销售商品，再按周转率升序、期末库存降序。",
        slow,
    )
    add_product_sheet(
        "Top热销品",
        "资生堂 2025年 Top 50 热销品",
        "按2025年净销售数量降序；保留平均库存、期末库存和周转率，便于在Excel内再次排序。",
        hot,
    )

    water_sheet = workbook.create_sheet("库存水位线")
    setup(
        water_sheet,
        "资生堂 2025年库存水位线",
        4,
        "按2025年每个月末历史库存快照展示；总库存=正装+小样，数量口径为件。",
    )
    headers(water_sheet, ["月份", "总库存", "正装库存", "小样库存"])
    water_rows = []
    for snapshot in SNAPSHOTS[1:]:
        formal_stock = stock_by_type["正装"][snapshot]
        sample_stock = stock_by_type["小样"][snapshot]
        water_rows.append(
            (snapshot.strftime("%Y-%m"), number(formal_stock + sample_stock), number(formal_stock), number(sample_stock))
        )
    add_rows(water_sheet, water_rows, int_columns=(2, 3, 4))
    finish(water_sheet, [13, 18, 18, 18], 16)
    chart = LineChart()
    chart.title = "2025年月末库存水位"
    chart.y_axis.title = "库存数量（件）"
    chart.x_axis.title = "月份"
    chart.style = 13
    chart.height = 10
    chart.width = 22
    chart.legend.position = "b"
    chart.add_data(Reference(water_sheet, min_col=2, max_col=4, min_row=4, max_row=16), titles_from_data=True)
    chart.set_categories(Reference(water_sheet, min_col=1, min_row=5, max_row=16))
    chart_colors = ("C65911", "ED7D31", "F4B183")
    for series, color in zip(chart.series, chart_colors):
        series.graphicalProperties.line.width = 24000
        series.graphicalProperties.line.solidFill = color
        series.marker.symbol = "circle"
        series.marker.size = 5
        series.marker.graphicalProperties.solidFill = color
        series.marker.graphicalProperties.line.solidFill = color
    water_sheet.add_chart(chart, "F4")

    workbook.properties.title = "资生堂_2025年周转分析"
    workbook.properties.creator = "Codex"
    workbook.save(output_path)

    check = load_workbook(output_path, read_only=False, data_only=False)
    expected_sheets = ["周转率汇总", "渠道周转率", "Top滞销品", "Top热销品", "库存水位线"]
    if check.sheetnames != expected_sheets or len(check["库存水位线"]._charts) != 1:
        raise RuntimeError("Excel结构校验失败")
    check.close()
    return {
        "sales_version": sales_version,
        "total_sales": total_sales,
        "average_stock": total_average_stock,
        "overall_rate": overall_rate,
        "formal_rate": sales_by_type["正装"] / average_stock["正装"],
        "sample_rate": sales_by_type["小样"] / average_stock["小样"],
        "channel_count": len(channels),
        "slow_count": len(slow),
        "hot_count": len(hot),
    }


if __name__ == "__main__":
    output_directory = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")
    os.makedirs(output_directory, exist_ok=True)
    output_file = os.path.join(output_directory, "资生堂_2025年周转分析.xlsx")
    result = build_workbook(output_file)
    print(f"FILE={output_file}")
    for key, value in result.items():
        print(f"{key}={value}")
