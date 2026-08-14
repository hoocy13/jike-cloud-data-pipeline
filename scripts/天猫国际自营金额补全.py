"""用项目内 Excel 的最后一列修正天猫国际自营销售单金额。"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import pymysql
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from config import DB_CONFIG


SOURCE_CHANNEL = "天猫国际自营"
DWD_TABLE = "销售单查询_进口超市上海仓补全"
MAPPING_TABLE = "天猫国际自营_销售单金额补全映射"
DEFAULT_XLSX = ROOT / "manual_inputs" / "天猫国际自营店数据补BI.xlsx"
JY_PATTERN = re.compile(r"JY\d+", re.IGNORECASE)


def db_config(database: str) -> dict[str, Any]:
    config = dict(DB_CONFIG)
    config["database"] = database
    return config


def decimal_value(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    raw = str(value).strip().replace(",", "")
    try:
        return Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"最后一列存在无法解析的金额: {value!r}") from exc


def build_mapping(xlsx_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"补数文件不存在: {xlsx_path}")

    workbook = load_workbook(xlsx_path, data_only=True, read_only=False)
    sheet = workbook.active
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    try:
        jy_column = headers.index("吉客云") + 1
    except ValueError as exc:
        raise ValueError(f"{xlsx_path} 缺少“吉客云”列") from exc
    amount_column = sheet.max_column
    if amount_column == jy_column:
        raise ValueError("最后一列不能与“吉客云”列相同")

    orders: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "source_rows": 0,
            "amount_rows": 0,
            "missing_rows": 0,
            "ambiguous_rows": [],
            "amount": Decimal("0"),
        }
    )
    positive_amount_without_jy = 0

    for row_number in range(2, sheet.max_row + 1):
        jy_text = str(sheet.cell(row_number, jy_column).value or "")
        order_ids = list(dict.fromkeys(match.upper() for match in JY_PATTERN.findall(jy_text)))
        amount = decimal_value(sheet.cell(row_number, amount_column).value)
        if not order_ids:
            if amount not in (None, Decimal("0")):
                positive_amount_without_jy += 1
            continue

        for order_id in order_ids:
            item = orders[order_id]
            item["source_rows"] += 1
            if len(order_ids) > 1:
                item["ambiguous_rows"].append(row_number)
            elif amount is None:
                item["missing_rows"] += 1
            else:
                item["amount_rows"] += 1
                item["amount"] += amount

    rows: list[dict[str, Any]] = []
    for order_id, item in sorted(orders.items()):
        if item["ambiguous_rows"]:
            status = "订单号歧义"
            corrected_amount = None
        elif item["missing_rows"]:
            status = "金额不完整"
            corrected_amount = None
        elif item["amount_rows"]:
            status = "完整匹配"
            corrected_amount = item["amount"].quantize(Decimal("0.01"))
        else:
            status = "缺少金额"
            corrected_amount = None
        rows.append(
            {
                "订单编号": order_id,
                "修正金额": corrected_amount,
                "来源行数": item["source_rows"],
                "金额行数": item["amount_rows"],
                "缺失金额行数": item["missing_rows"],
                "歧义行数": len(item["ambiguous_rows"]),
                "匹配状态": status,
                "歧义Excel行": ",".join(str(value) for value in item["ambiguous_rows"]),
                "线下映射文件": xlsx_path.name,
            }
        )

    workbook.close()
    summary = {
        "sheet": sheet.title,
        "amount_column": amount_column,
        "amount_header": headers[amount_column - 1],
        "excel_rows": sheet.max_row - 1,
        "orders": len(rows),
        "positive_amount_rows_without_jy": positive_amount_without_jy,
        "status": {
            status: sum(row["匹配状态"] == status for row in rows)
            for status in ("完整匹配", "金额不完整", "订单号歧义", "缺少金额")
        },
    }
    return rows, summary


def ensure_schema(cursor: Any) -> None:
    cursor.execute("CREATE DATABASE IF NOT EXISTS `dwd` DEFAULT CHARACTER SET utf8mb4")


def publish_mapping(rows: list[dict[str, Any]], apply: bool) -> dict[str, Any]:
    connection = pymysql.connect(**db_config("dwd"))
    stage = f"{MAPPING_TABLE}_stage"
    old = f"{MAPPING_TABLE}_old"
    try:
        with connection.cursor() as cursor:
            ensure_schema(cursor)
            cursor.execute(f"DROP TABLE IF EXISTS `dwd`.`{stage}`")
            cursor.execute(
                f"""
                CREATE TABLE `dwd`.`{stage}` (
                  `订单编号` VARCHAR(100) NOT NULL,
                  `修正金额` DECIMAL(20,2) NULL,
                  `来源行数` INT NOT NULL,
                  `金额行数` INT NOT NULL,
                  `缺失金额行数` INT NOT NULL,
                  `歧义行数` INT NOT NULL,
                  `匹配状态` VARCHAR(20) NOT NULL,
                  `歧义Excel行` TEXT NULL,
                  `线下映射文件` VARCHAR(255) NOT NULL,
                  `updatetime` DATETIME NOT NULL,
                  PRIMARY KEY (`订单编号`),
                  KEY `idx_匹配状态` (`匹配状态`)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """
            )
            now = datetime.now()
            cursor.executemany(
                f"""
                INSERT INTO `dwd`.`{stage}`
                  (`订单编号`,`修正金额`,`来源行数`,`金额行数`,`缺失金额行数`,
                   `歧义行数`,`匹配状态`,`歧义Excel行`,`线下映射文件`,`updatetime`)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                [
                    (
                        row["订单编号"],
                        row["修正金额"],
                        row["来源行数"],
                        row["金额行数"],
                        row["缺失金额行数"],
                        row["歧义行数"],
                        row["匹配状态"],
                        row["歧义Excel行"] or None,
                        row["线下映射文件"],
                        now,
                    )
                    for row in rows
                ],
            )
            cursor.execute(f"DROP TABLE IF EXISTS `dwd`.`{old}`")
            cursor.execute(
                "SELECT TABLE_TYPE FROM information_schema.tables "
                "WHERE table_schema='dwd' AND table_name=%s",
                (MAPPING_TABLE,),
            )
            if cursor.fetchone():
                cursor.execute(
                    f"RENAME TABLE `dwd`.`{MAPPING_TABLE}` TO `dwd`.`{old}`, "
                    f"`dwd`.`{stage}` TO `dwd`.`{MAPPING_TABLE}`"
                )
                cursor.execute(f"DROP TABLE `dwd`.`{old}`")
            else:
                cursor.execute(f"RENAME TABLE `dwd`.`{stage}` TO `dwd`.`{MAPPING_TABLE}`")

            reset_rows = 0
            updated_rows = 0
            if apply:
                cursor.execute(
                    f"""
                    UPDATE `dwd`.`{DWD_TABLE}` s
                    JOIN `ods`.`销售单查询` o
                      ON o.`订单编号` = s.`订单编号`
                     AND o.`销售渠道` = s.`销售渠道`
                    SET s.`应收合计` = o.`应收合计`,
                        s.`实付金额` = o.`实付金额`
                    WHERE s.`销售渠道` = %s
                    """,
                    (SOURCE_CHANNEL,),
                )
                reset_rows = int(cursor.rowcount)
                cursor.execute(
                    f"""
                    UPDATE `dwd`.`{DWD_TABLE}` s
                    JOIN `dwd`.`{MAPPING_TABLE}` m ON m.`订单编号` = s.`订单编号`
                    SET s.`应收合计` = m.`修正金额`,
                        s.`实付金额` = m.`修正金额`
                    WHERE s.`销售渠道` = %s
                      AND m.`匹配状态` = '完整匹配'
                      AND m.`修正金额` IS NOT NULL
                    """,
                    (SOURCE_CHANNEL,),
                )
                updated_rows = int(cursor.rowcount)

            cursor.execute(
                f"""
                SELECT COUNT(*),
                       SUM(m.`匹配状态` = '完整匹配'),
                       SUM(m.`匹配状态` = '完整匹配' AND s.`订单编号` IS NOT NULL)
                FROM `dwd`.`{MAPPING_TABLE}` m
                LEFT JOIN `dwd`.`{DWD_TABLE}` s
                  ON s.`订单编号` = m.`订单编号`
                 AND s.`销售渠道` = %s
                """,
                (SOURCE_CHANNEL,),
            )
            mapping_rows, complete_rows, complete_in_sales = [
                int(value or 0) for value in cursor.fetchone()
            ]
        connection.commit()
        return {
            "mapping_rows": mapping_rows,
            "complete_rows": complete_rows,
            "complete_in_sales": complete_in_sales,
            "reset_rows": reset_rows,
            "updated_rows": updated_rows,
            "applied": apply,
        }
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按项目内 Excel 最后一列修正天猫国际自营销售单金额")
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"补数 Excel；默认 {DEFAULT_XLSX}",
    )
    parser.add_argument("--apply", action="store_true", help="将完整匹配金额写入 DWD；默认只刷新映射表")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows, excel_summary = build_mapping(args.xlsx.resolve())
    database_summary = publish_mapping(rows, args.apply)
    print(
        "[DONE] "
        + json.dumps(
            {"excel": excel_summary, "database": database_summary},
            ensure_ascii=False,
            default=str,
        ),
        flush=True,
    )


if __name__ == "__main__":
    main()
